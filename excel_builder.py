"""Build 26-column webshop import Excel matching Eszter's template."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Exact headers from "új termék kreálás_2025_szept.xlsx"
HEADERS = [
    "Cikkszám",                                          # A
    "Kategória azonosító(k)",                             # B
    "Terméknév (hu)",                                     # C
    "Rövid leírás (hu)",                                  # D
    "Hosszú leírás (hu)",                                 # E
    "További termékképek",                                # F
    "Mikortól kapható",                                   # G
    "Gyártói cikkszám",                                   # H
    "Elsődleges termékkép",                               # I
    "Szállítandó termék(igen (1) v. nem (0))",            # J
    "Státusz (engedélyezett (1) v. letiltott (0) v. kifutott (2))",  # K
    "Ingyenesen szállítható (Igen / Nem)",                # L
    "Bruttó ár",                                          # M
    "Alapár",                                             # N
    "Gyártó",                                             # O
    "Nincs készleten állapot",                            # P
    "Csak Raktár 1 készleten állapot",                    # Q
    "Csak Raktár 2 készleten állapot",                    # R
    "Raktárkészlet 1",                                    # S
    "Terméktípus",                                        # T
    "Tulajdonság: Szín (szin)",                           # U
    "Tulajdonság: Anyag (anyag)",                         # V
    "Tulajdonság: Szélesség (szelesseg)",                 # W
    "Tulajdonság: Hosszúság (hosszusag)",                 # X
    "Tulajdonság: Magasság (magassag)",                   # Y
    "Tulajdonság: Ajánlott (ajanlott)",                   # Z
]

COL_WIDTHS = [12, 18, 35, 50, 60, 40, 14, 16, 40, 8, 8, 10, 12, 12, 16, 16, 16, 16, 10, 14, 12, 20, 12, 12, 12, 10]

HEADER_FONT = Font(bold=True, size=10)


def build_catalogue_excel(products: list[dict]) -> bytes:
    """Build a 26-column webshop import Excel.

    Each product dict should have:
        - eszter_sku: her SKU (e.g. HN26-7)
        - manufacturer_sku: HN SKU (e.g. 1001362)
        - name_hu: Hungarian product name
        - short_desc: short description
        - long_desc: long description (HTML)
        - product_type: from fixed list
        - color_hu: Hungarian color
        - material: material string
        - width, length, height: dimensions
        - primary_image: main image URL
        - additional_images: list of additional image URLs
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "export"

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    today = datetime.now().strftime("%Y-%m-%d")

    for i, p in enumerate(products):
        row = i + 2

        # Additional images joined with |||
        additional_imgs = p.get("additional_images", [])
        additional_imgs_str = "|||".join(additional_imgs) if additional_imgs else ""

        ws.cell(row=row, column=1, value=p.get("eszter_sku", ""))         # A: Cikkszám
        ws.cell(row=row, column=2, value="")                               # B: Kategória (manual)
        ws.cell(row=row, column=3, value=p.get("name_hu", ""))            # C: Terméknév
        ws.cell(row=row, column=4, value=p.get("short_desc", ""))         # D: Rövid leírás
        ws.cell(row=row, column=5, value=p.get("long_desc", ""))          # E: Hosszú leírás
        ws.cell(row=row, column=6, value=additional_imgs_str)              # F: További képek
        ws.cell(row=row, column=7, value=today)                            # G: Mikortól kapható
        ws.cell(row=row, column=8, value=p.get("manufacturer_sku", ""))   # H: Gyártói cikkszám
        ws.cell(row=row, column=9, value=p.get("primary_image", ""))      # I: Elsődleges kép
        ws.cell(row=row, column=10, value=1)                               # J: Szállítandó
        ws.cell(row=row, column=11, value=1)                               # K: Státusz
        ws.cell(row=row, column=12, value="Igen")                          # L: Ingyenesen szállítható
        ws.cell(row=row, column=13, value="")                              # M: Bruttó ár (manual)
        ws.cell(row=row, column=14, value="")                              # N: Alapár (manual)
        ws.cell(row=row, column=15, value="House Nordic")                  # O: Gyártó
        ws.cell(row=row, column=16, value="Elfogyott")                     # P: Nincs készleten
        ws.cell(row=row, column=17, value="Raktáron")                      # Q: Raktár 1
        ws.cell(row=row, column=18, value="Előrendelhető")                 # R: Raktár 2
        ws.cell(row=row, column=19, value=10)                              # S: Raktárkészlet
        ws.cell(row=row, column=20, value=p.get("product_type", ""))      # T: Terméktípus
        ws.cell(row=row, column=21, value=p.get("color_hu", ""))          # U: Szín
        ws.cell(row=row, column=22, value=p.get("material", ""))          # V: Anyag
        ws.cell(row=row, column=23, value=p.get("width", ""))             # W: Szélesség
        ws.cell(row=row, column=24, value=p.get("length", ""))            # X: Hosszúság
        ws.cell(row=row, column=25, value=p.get("height", ""))            # Y: Magasság
        ws.cell(row=row, column=26, value="")                              # Z: Ajánlott

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
