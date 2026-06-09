"""Nordic Home Product Catalogue Builder — Streamlit Web App."""

import os

import openpyxl
import streamlit as st
from dotenv import load_dotenv

from hn_lookup import fetch_product, download_product_image
from extractor import generate_hu_content
from excel_builder import build_catalogue_excel

load_dotenv()

# --- Page Config ---
st.set_page_config(page_title="Nordic Home Catalogue Builder", page_icon="🏠", layout="wide")
st.title("🏠 Nordic Home Catalogue Builder")
st.caption("Upload your House Nordic product list → generate Hungarian descriptions → download webshop import Excel")

# --- Session state ---
if "processed_products" not in st.session_state:
    st.session_state.processed_products = []
if "processing_done" not in st.session_state:
    st.session_state.processing_done = False

# --- Sidebar ---
with st.sidebar:
    st.header("Settings")
    api_key = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-...")
    if api_key:
        os.environ["ANTHROPIC_API_KEY"] = api_key

    st.divider()
    st.metric("Products processed", len(st.session_state.processed_products))

    if st.button("🗑️ Clear session"):
        st.session_state.processed_products = []
        st.session_state.processing_done = False
        st.rerun()

# --- Main Area ---
uploaded_file = st.file_uploader(
    "Upload HouseNordic product Excel",
    type=["xlsx"],
    help="Excel with columns: Cikkszám, Gyártói cikkszám, Terméknév",
)

if uploaded_file:
    # Parse uploaded Excel
    wb = openpyxl.load_workbook(uploaded_file, read_only=True)
    ws = wb.active

    input_products = []
    for row in range(2, ws.max_row + 1):
        eszter_sku = ws.cell(row, 1).value
        manufacturer_sku = ws.cell(row, 2).value
        en_name = ws.cell(row, 3).value
        if eszter_sku and manufacturer_sku:
            input_products.append({
                "eszter_sku": str(eszter_sku).strip(),
                "manufacturer_sku": str(int(manufacturer_sku) if isinstance(manufacturer_sku, float) else manufacturer_sku).strip(),
                "en_name": str(en_name or "").strip(),
            })
    wb.close()

    st.success(f"Loaded **{len(input_products)}** products from uploaded file")

    # Preview input
    with st.expander("Preview uploaded products", expanded=False):
        preview = [{"SKU": p["eszter_sku"], "Manufacturer SKU": p["manufacturer_sku"], "Name": p["en_name"]} for p in input_products[:10]]
        st.dataframe(preview, use_container_width=True)
        if len(input_products) > 10:
            st.caption(f"... and {len(input_products) - 10} more")

    # Process button
    if st.button("🚀 Process All Products", type="primary", disabled=not os.environ.get("ANTHROPIC_API_KEY")):
        if not os.environ.get("ANTHROPIC_API_KEY"):
            st.error("Paste your Anthropic API Key in the sidebar first.")
            st.stop()

        st.session_state.processed_products = []
        st.session_state.processing_done = False

        progress_bar = st.progress(0, text="Starting...")
        status_container = st.empty()

        for i, inp in enumerate(input_products):
            sku = inp["manufacturer_sku"]
            progress_bar.progress(
                (i + 1) / len(input_products),
                text=f"Processing {i+1}/{len(input_products)}: {inp['eszter_sku']} ({sku})",
            )

            # Step 1: Fetch product data from House Nordic
            product_data = fetch_product(sku)
            if not product_data:
                status_container.warning(f"⚠️ SKU {sku} not found on housenordic.dk — skipping")
                continue

            # Step 2: Download product image for AI analysis
            image_bytes = download_product_image(product_data.get("primary_image", ""))

            # Step 3: Generate Hungarian content with Claude
            try:
                hu_content = generate_hu_content(product_data, image_bytes)
            except Exception as e:
                status_container.warning(f"⚠️ AI generation failed for {sku}: {e}")
                hu_content = {
                    "name_hu": product_data.get("title", inp["en_name"]),
                    "short_desc": "",
                    "long_desc": "",
                    "product_type": "",
                    "color_hu": "",
                }

            # Step 4: Build output row
            images = product_data.get("images", [])
            primary_image = images[0] if images else product_data.get("primary_image", "")
            additional_images = images[1:] if len(images) > 1 else []

            processed = {
                "eszter_sku": inp["eszter_sku"],
                "manufacturer_sku": sku,
                "name_hu": hu_content["name_hu"],
                "short_desc": hu_content["short_desc"],
                "long_desc": hu_content["long_desc"],
                "product_type": hu_content["product_type"],
                "color_hu": hu_content["color_hu"],
                "material": product_data.get("material_da") or product_data.get("material_en", ""),
                "width": product_data.get("width", ""),
                "length": product_data.get("length", ""),
                "height": product_data.get("height", ""),
                "primary_image": primary_image,
                "additional_images": additional_images,
            }
            st.session_state.processed_products.append(processed)

        progress_bar.progress(1.0, text="Done!")
        st.session_state.processing_done = True
        st.rerun()

# --- Results ---
if st.session_state.processed_products:
    st.divider()
    st.subheader(f"Processed Products ({len(st.session_state.processed_products)})")

    display_data = [
        {
            "SKU": p["eszter_sku"],
            "Terméknév": p["name_hu"],
            "Típus": p["product_type"],
            "Szín": p["color_hu"],
            "Méret": f"{p['width']}x{p['length']}x{p['height']}",
        }
        for p in st.session_state.processed_products
    ]
    st.dataframe(display_data, use_container_width=True)

    # Expandable detail view
    with st.expander("View full descriptions", expanded=False):
        for p in st.session_state.processed_products[:5]:
            st.markdown(f"**{p['eszter_sku']} — {p['name_hu']}**")
            st.text(f"Rövid: {p['short_desc']}")
            st.markdown(f"Hosszú: {p['long_desc']}")
            st.divider()

    # Download
    excel_bytes = build_catalogue_excel(st.session_state.processed_products)
    st.download_button(
        "⬇️ Download webshop import Excel (.xlsx)",
        data=excel_bytes,
        file_name="uj_termek_krealas_HouseNordic.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
